import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook 
import os
from datetime import datetime, timedelta
import re 
import numpy as np 

# =========================================================================
# 1. 設定とXPath (設定とXPath)
# =========================================================================

VNDIRECT_URL = "https://banggia.vndirect.com.vn/chung-khoan/hose"
EXCEL_FILE_NAME = "VNDirect_data.xlsx"
TIMEOUT = 20
# Chromeプロファイルディレクトリのパス (ご自身の環境に合わせて変更してください)
USER_DATA_DIR = r"C:\Users\A22M\Programming\Python\Chrome VPS Profile" 

XPATH_SELECTORS = {
    "VNIndex":'//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[1]/span[3]', 
    "Spread_Icon":'//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[1]/span[2]', # XPath: 上昇/下降アイコン
    "Spread":'//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[1]/span[4]', # XPath: Spread と Spread% の両方の値を取得
    "Value":'//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[2]/span[3]', 
    "Volume":'//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[2]/span[1]',
    "Meigara_Up":'//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[3]/span[2]', # 上昇銘柄数
    "Meigara_Down":'//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[3]/span[7]', # 下落銘柄数
    "Meigara_Unchanged":'//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[3]/span[5]', # 不変銘柄数
}

FINAL_COLUMN_ORDER = [
    'Date',
    'VNIndex',
    'Spread',
    'Spread%', 
    'Value', 
    'Volume', 
    'Meigara_Up', 
    'Meigara_Down', 
    'Meigara_Unchanged',
]
# 比較に使用する列 ( 'Date' を除く)
COMPARE_COLUMNS = [col for col in FINAL_COLUMN_ORDER if col != 'Date']

# Excel読み込み時に比較列をすべて文字列(str)に強制キャストし、混在型エラーを回避
DTYPE_CONVERTERS = {col: str for col in COMPARE_COLUMNS}

# =========================================================================
# 2. データチェックと取引日の決定を支援する関数
# =========================================================================

def get_trading_date():
    """現在の時刻 (午前9時前/後) に基づいて取引日を決定します。"""
    
    now = datetime.now()
    # 当日の午前9時を設定
    opening_time = now.replace(hour=9, minute=0, second=0, microsecond=0)
    
    # 0 = 月曜日, 6 = 日曜日
    weekday = now.weekday() 
    
    if now < opening_time or weekday >= 5: # もし午前9時前、または土日であれば
        # 最終取引日を見つける必要がある: 月曜〜金曜に当たるまで日付を遡る
        current_date = now.date()
        
        # 1日ずつ遡り始める
        while True:
            current_date -= timedelta(days=1)
            trading_weekday = current_date.weekday()
            
            # 有効な取引日(月〜金)であれば、この日付を使用
            if trading_weekday >= 0 and trading_weekday <= 4: 
                return current_date.strftime("%d/%m/%Y")
            
    else:
        # 午前9時以降かつ月〜金であれば、当日を使用
        return now.strftime("%d/%m/%Y")


def normalize_value_for_comparison(value):
    """値を比較用の標準文字列形式に変換します。"""
    if value is None or (isinstance(value, (float, np.number)) and np.isnan(value)):
        return "N/A"
    
    # 文字列の場合、不要な文字を削除
    if isinstance(value, str):
        # 書式設定文字を削除
        return value.strip().replace(',', '')
    
    # 数値 (float/int) の場合、適切な精度で文字列に変換
    try:
        if isinstance(value, (float, int)):
            if value.is_integer():
                return str(int(value))
            # 小数点以下3桁にフォーマット
            return "{:.3f}".format(value)
    except:
        pass 
        
    return str(value).strip().replace(',', '')


def get_last_excel_data():
    """Excelファイルから最終行のデータ (標準化済み) を読み込んで返します。"""
    if not os.path.isfile(EXCEL_FILE_NAME):
        return None
    try:
        # 比較に必要な列のみを読み込み ('Date' を除外) し、型を強制的にstrにする
        df = pd.read_excel(EXCEL_FILE_NAME, usecols=COMPARE_COLUMNS, dtype=DTYPE_CONVERTERS) 
        
        if df.empty:
            return None
            
        last_row = df.iloc[-1].to_dict()
        normalized_data = {}
        
        # 比較列を標準化
        for col in COMPARE_COLUMNS:
            normalized_data[col] = normalize_value_for_comparison(last_row.get(col))
            
        return normalized_data
        
    except Exception as e:
        print(f"⚠️ 警告: Excelファイルの最終行の読み込みと標準化中にエラーが発生しました: {e}。重複チェックをスキップします。")
        return None

# =========================================================================
# 3. メインデータ取得関数 (日付ロジック更新済み)
# =========================================================================

def get_market_data_and_save():
    print("🚀 仮想ブラウザを起動しています...")
    chrome_options = Options()
    chrome_options.add_argument(f"user-data-dir={USER_DATA_DIR}") 
    chrome_options.add_argument("--window-size=1920,1080")

    driver = None
    try:
        driver = webdriver.Chrome(options=chrome_options)
    except Exception as e:
        print(f"❌ WebDriverの初期化エラー: {e}")
        return

    data_row = {key: "N/A" for key in COMPARE_COLUMNS}
    is_spread_negative = False 

    try:
        print(f"🌐 ウェブサイトにアクセス中: {VNDIRECT_URL}")
        driver.get(VNDIRECT_URL)

        WebDriverWait(driver, TIMEOUT).until(
            EC.presence_of_element_located((By.XPATH, XPATH_SELECTORS['VNIndex']))
        )
        print("✅ VNIndexの準備ができました。")

        # --- ステップ 1: Spreadアイコンに基づいて上昇/下降傾向を決定 ---
        try:
            icon_element = driver.find_element(By.XPATH, XPATH_SELECTORS['Spread_Icon'])
            icon_class = icon_element.get_attribute("class")
            
            if "icon-arrowdown" in icon_class.lower():
                is_spread_negative = True
                print("⬇️ Spread傾向: 下降 ('-'記号を追加します)。")
            else:
                is_spread_negative = False
                print("⬆️ Spread傾向: 上昇/不変 (そのまま保持します)。")
                
        except Exception as e:
            # アイコンが見つからない場合のエラー処理
            error_msg = str(e).split('\n')[0].replace('Message: ', '')
            print(f"⚠️ 警告: Spreadアイコンが見つかりませんでした ({error_msg})。デフォルトでSpreadは上昇とします。")


        # --- ステップ 2: データの取得と処理 (Spreadロジックの適用) ---
        for name, selector in XPATH_SELECTORS.items():
            if name == "Spread_Icon":
                continue 

            try:
                element = driver.find_element(By.XPATH, selector) 
                value = element.text.strip()
                
                # *** Spread と Spread% の処理ロジック ***
                if name == "Spread":
                    raw_spread = "N/A"
                    raw_spread_percent = "N/A"
                    
                    # 例: "16.55 1.55%" のような形式を検索
                    match = re.search(r'([\d\.\,\-]+)\s+([\d\.\,\-]+%)', value)
                    
                    if match:
                        raw_spread = match.group(1).strip().replace(',', '')
                        raw_spread_percent = match.group(2).strip().replace('%', '') 
                    
                    elif '/' in value:
                         # 例: "16.55/1.55%" のような形式を処理
                         parts = value.split('/')
                         raw_spread = parts[0].strip().replace(',', '')
                         raw_spread_percent = parts[1].strip().replace('%', '')

                    # 下降傾向の場合、マイナス記号を適用 (両方の列に)
                    if is_spread_negative:
                        if raw_spread != "N/A" and not raw_spread.startswith('-'):
                            data_row['Spread'] = "-" + raw_spread
                        else:
                            data_row['Spread'] = raw_spread
                        
                        if raw_spread_percent != "N/A" and not raw_spread_percent.startswith('-'):
                            data_row['Spread%'] = "-" + raw_spread_percent
                        else:
                            data_row['Spread%'] = raw_spread_percent

                    else:
                        data_row['Spread'] = raw_spread
                        data_row['Spread%'] = raw_spread_percent
                        
                    print(f"   -> Spread (ポイント): {data_row['Spread']}")
                    print(f"   -> Spread% (指標): {data_row['Spread%']}")
                    continue 
                # ***********************************

                # *** Value (桁区切り形式) の処理ロジック ***
                if name == "Value":
                    temp_value = value.replace(' tỷ', '').strip() # ' tỷ' (ビリオン) を削除
                    temp_value = temp_value.replace(',', '')
                    match_final = re.search(r'([\d.]+)', temp_value)
                    
                    if match_final:
                        raw_number_str = match_final.group(1)
                        try:
                            num_value = float(raw_number_str)
                            # カンマ区切り形式にフォーマット
                            value = "{:,.3f}".format(num_value)
                        except ValueError:
                            value = raw_number_str
                    else:
                        value = "N/A"
                # ***********************************
                
                # 3. その他の指標のデータを更新 (Meigara_Up, Meigara_Down, Meigara_Unchangedを含む)
                data_row[name] = value 
                if name != "Spread": 
                    print(f"   -> {name}: {value}")
            
            except Exception as e:
                # XPATH_SELECTORSキーがそのまま変数名として使われるため、エラーメッセージを調整
                if name in ["Meigara_Up", "Meigara_Down", "Meigara_Unchanged"]:
                    print(f"❌ エラー: 要素 {name} が見つかりません | 詳細: {str(e).split('\n')[0].replace('Message: ', '')}")
                else:
                    error_msg = str(e).split('\n')[0].replace('Message: ', '')
                    print(f"❌ エラー: 要素 {name} が見つかりません | 詳細: {error_msg}")

                data_row[name] = "N/A" 
                if name == "Spread":
                    data_row['Spread%'] = "N/A"

        # --- ステップ 3: 重複チェックとファイル書き込み ---
        
        # 3a. Excelファイルから最終行のデータ (標準化済み) を取得
        last_data_normalized = get_last_excel_data() 
        
        # 取得したデータを比較用に標準化
        current_data_normalized = {col: normalize_value_for_comparison(data_row.get(col)) for col in COMPARE_COLUMNS}
        
        is_duplicate = False
        if last_data_normalized:
            # 標準化された値を比較
            is_duplicate = all(current_data_normalized.get(col) == last_data_normalized.get(col) for col in COMPARE_COLUMNS)

        if is_duplicate:
            # 通知 1: 重複データ
            print("\n=======================================================")
            print("🚫 現在のデータはExcelの最終データと完全に同じです。")
            print("➡️ **既存のデータが最新のデータです** (取引セッションは終了した可能性があります)。")
            print("=======================================================\n")
            # データが重複している場合、関数を終了
            return 
        
        # 通知 2: 新しいデータ (ファイル書き込みに進む)
        print("\n=======================================================")
        print("✅ 新しいデータが収集されました!")
        print("➡️ **今回収集されたデータが最新のデータです**。")
        print("=======================================================\n")
        
        # 3b. データが重複していない場合、Excelファイルに書き込む
        
        # 調整された取引日を取得し設定
        data_row['Date'] = get_trading_date() 
        
        df = pd.DataFrame([data_row])[FINAL_COLUMN_ORDER]

        print(f"💾 データを {EXCEL_FILE_NAME} に書き込んでいます")
        file_exists = os.path.isfile(EXCEL_FILE_NAME)
        
        if file_exists:
            try:
                book = load_workbook(EXCEL_FILE_NAME)
                # 既存のファイルに追記
                with pd.ExcelWriter(EXCEL_FILE_NAME, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    sheet = writer.book.active 
                    start_row = sheet.max_row
                    df.to_excel(writer, sheet_name=sheet.title, startrow=start_row, index=False, header=False)
            except Exception as e:
                print(f"⚠️ 警告: データ追記エラー ({e})、ファイルを上書きします。")
                # 追記失敗時はヘッダー付きで上書き
                df.to_excel(EXCEL_FILE_NAME, index=False, header=True, engine='openpyxl')
        else:
            # ファイルが存在しない場合は新規作成
            df.to_excel(EXCEL_FILE_NAME, index=False, header=True, engine='openpyxl')

        print("🎉 ファイル書き込み完了!")

    except Exception as e:
        print(f"❌ 全体的なデータスクレイピングエラー: {e}")

    finally:
        if driver:
            driver.quit()
            print("🔒 ブラウザを閉じます。")

# =========================================================================
# 4. メイン処理
# =========================================================================

if __name__ == "__main__":
    get_market_data_and_save()