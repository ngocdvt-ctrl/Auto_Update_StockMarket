import os
import re
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# =========================================================================
# 0. PATH / BASE DIR
# =========================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

VNDIRECT_URL = "https://banggia.vndirect.com.vn/chung-khoan/hose"
EXCEL_FILE_NAME = "VNDirect_data.xlsx"
EXCEL_FILE_PATH = os.path.join(BASE_DIR, EXCEL_FILE_NAME)

TIMEOUT = 20

# Chrome ユーザープロファイル（必要に応じて変更）
USER_DATA_DIR = r"C:\Users\A22M\Programming\Python\Chrome VPS Profile"


# =========================================================================
# 1. 取得する要素（XPATH）
# =========================================================================

XPATH_SELECTORS = {
    "VNIndex": '//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[1]/span[3]',
    "Spread_Icon": '//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[1]/span[2]',  # 上下矢印アイコン
    "Spread": '//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[1]/span[4]',       # Spread と Spread% が入る
    "Value": '//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[2]/span[3]',
    "Volume": '//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[2]/span[1]',
    "CP_Tang": '//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[3]/span[2]',
    "CP_Giam": '//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[3]/span[7]',
    "CP_KhongDoi": '//*[@id="charts-wrapper"]/div/div/div[1]/div[2]/p[3]/span[5]',
}


# =========================================================================
# 2. カラム名（日本語）定義
# =========================================================================

# 内部キー（英語） -> Excel出力用（日本語）
COLUMN_JP = {
    "ThoiGian": "取引日",
    "VNIndex": "VN指数",
    "Spread": "前日比(ポイント)",
    "Spread%": "前日比(%)",
    "Value": "売買代金(億VND)",  # webの "tỷ" (=10^9 VND) 表示をそのまま数値で保存
    "Volume": "出来高",
    "CP_Tang": "上昇銘柄数",
    "CP_Giam": "下落銘柄数",
    "CP_KhongDoi": "変わらず銘柄数",
}

FINAL_COLUMN_ORDER_INTERNAL = [
    "ThoiGian",
    "VNIndex",
    "Spread",
    "Spread%",
    "Value",
    "Volume",
    "CP_Tang",
    "CP_Giam",
    "CP_KhongDoi",
]
FINAL_COLUMN_ORDER_JP = [COLUMN_JP[c] for c in FINAL_COLUMN_ORDER_INTERNAL]

# 比較用（取引日は除外）
COMPARE_COLUMNS_INTERNAL = [c for c in FINAL_COLUMN_ORDER_INTERNAL if c != "ThoiGian"]
COMPARE_COLUMNS_JP = [COLUMN_JP[c] for c in COMPARE_COLUMNS_INTERNAL]

# ログ表示用（内部キー -> 日本語ラベル）
LOG_LABEL = {
    "VNIndex": "VN指数",
    "Spread": "前日比(ポイント)",
    "Spread%": "前日比(%)",
    "Value": "売買代金(億VND)",
    "Volume": "出来高",
    "CP_Tang": "上昇銘柄数",
    "CP_Giam": "下落銘柄数",
    "CP_KhongDoi": "変わらず銘柄数",
}

# 型（numericで保存するため）
TYPE_MAP_INTERNAL = {
    "VNIndex": float,
    "Spread": float,
    "Spread%": float,
    "Value": float,
    "Volume": int,
    "CP_Tang": int,
    "CP_Giam": int,
    "CP_KhongDoi": int,
}


# =========================================================================
# 3. 補助関数：取引日判定、パース、Excel最後行取得、比較
# =========================================================================

def get_trading_date() -> str:
    """現在時刻に基づいて取引日を判定する（9:00 前 / 土日なら直近営業日）。"""
    now = datetime.now()
    opening_time = now.replace(hour=9, minute=0, second=0, microsecond=0)
    weekday = now.weekday()  # 0=Mon ... 6=Sun

    if now < opening_time or weekday >= 5:
        current_date = now.date()
        while True:
            current_date -= timedelta(days=1)
            if 0 <= current_date.weekday() <= 4:
                return current_date.strftime("%d/%m/%Y")
    else:
        return now.strftime("%d/%m/%Y")


def _clean_number_text(s: str) -> str:
    """Remove common separators/spaces."""
    if s is None:
        return ""
    return str(s).strip().replace(",", "").replace("\u00a0", " ")


def parse_float(text) -> float | None:
    """Parse float from text like '1,234.56' or '+12.3' or '0.56%'."""
    if text is None:
        return None
    s = _clean_number_text(text)
    s = s.replace("%", "").strip()
    m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def parse_int(text) -> int | None:
    """Parse int from text like '1,234,567'."""
    if text is None:
        return None
    s = _clean_number_text(text)
    m = re.search(r"\d+", s)
    if not m:
        return None
    try:
        return int(m.group(0))
    except Exception:
        return None


def parse_value_ty(text) -> float | None:
    """
    Parse 'Value' that appears like '12.345 tỷ' => 12.345 (float).
    NOTE:
      - This keeps the number in "tỷ" units (10^9 VND) as a float.
      - If you want VND, multiply by 1_000_000_000.
    """
    if text is None:
        return None
    s = str(text).replace("tỷ", "").replace(" tỷ", "")
    s = _clean_number_text(s)
    m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def get_last_excel_data_numeric() -> dict | None:
    """Excelの最終行（比較対象カラム）を読み、numericに揃えて返す。"""
    if not os.path.isfile(EXCEL_FILE_PATH):
        return None

    try:
        df = pd.read_excel(EXCEL_FILE_PATH, usecols=COMPARE_COLUMNS_JP)
        if df.empty:
            return None

        last_row = df.iloc[-1].to_dict()

        # Convert to numeric using schema
        out = {}
        for internal_key in COMPARE_COLUMNS_INTERNAL:
            jp_col = COLUMN_JP[internal_key]
            target_type = TYPE_MAP_INTERNAL.get(internal_key, str)

            v = last_row.get(jp_col, None)
            if v is None or (isinstance(v, float) and np.isnan(v)):
                out[jp_col] = None
                continue

            if target_type is float:
                # pandas may already read as float; if string then parse
                if isinstance(v, (int, float, np.number)):
                    out[jp_col] = float(v)
                else:
                    out[jp_col] = parse_float(v)
            elif target_type is int:
                if isinstance(v, (int, np.integer)):
                    out[jp_col] = int(v)
                elif isinstance(v, (float, np.floating)) and not np.isnan(v):
                    # excel sometimes stores ints as floats
                    out[jp_col] = int(v)
                else:
                    out[jp_col] = parse_int(v)
            else:
                out[jp_col] = str(v)

        return out

    except Exception as e:
        print(f"⚠️ Excel 読み込みでエラー: {e}。重複チェックをスキップします。")
        return None


def is_duplicate_numeric(current_jp: dict, last_jp: dict, float_tol: float = 1e-6) -> bool:
    """Numeric comparison with tolerance for floats."""
    if not last_jp:
        return False

    for internal_key in COMPARE_COLUMNS_INTERNAL:
        jp_col = COLUMN_JP[internal_key]
        target_type = TYPE_MAP_INTERNAL.get(internal_key, str)

        cur = current_jp.get(jp_col, None)
        last = last_jp.get(jp_col, None)

        # Treat both None as equal
        if cur is None and last is None:
            continue
        if cur is None or last is None:
            return False

        if target_type is float:
            try:
                if abs(float(cur) - float(last)) > float_tol:
                    return False
            except Exception:
                return False
        elif target_type is int:
            try:
                if int(cur) != int(last):
                    return False
            except Exception:
                return False
        else:
            if str(cur) != str(last):
                return False

    return True


# =========================================================================
# 4. メイン処理：取得 + 重複チェック + Excel追記
# =========================================================================

def get_market_data_and_save():
    # ログ：保存先の確認
    print("📌 実行ディレクトリ(CWD):", os.getcwd())
    print("📌 スクリプトのフォルダ:", BASE_DIR)
    print("📌 Excel 保存パス:", EXCEL_FILE_PATH)

    print("\n🚀 ブラウザを起動中...")
    chrome_options = Options()
    chrome_options.add_argument(f"user-data-dir={USER_DATA_DIR}")
    chrome_options.add_argument("--window-size=1920,1080")

    driver = None
    try:
        driver = webdriver.Chrome(options=chrome_options)
    except Exception as e:
        print(f"❌ WebDriver 初期化エラー: {e}")
        return

    # 取得データ（内部キー）※ numericで持つ
    data_row_internal = {key: None for key in COMPARE_COLUMNS_INTERNAL}
    is_spread_negative = False

    try:
        print(f"🌐 サイトへアクセス: {VNDIRECT_URL}")
        driver.get(VNDIRECT_URL)

        WebDriverWait(driver, TIMEOUT).until(
            EC.presence_of_element_located((By.XPATH, XPATH_SELECTORS["VNIndex"]))
        )
        print("✅ VN指数の要素を検出しました。")

        # --- 1) Spread の増減方向をアイコンで判定 ---
        try:
            icon_element = driver.find_element(By.XPATH, XPATH_SELECTORS["Spread_Icon"])
            icon_class = (icon_element.get_attribute("class") or "").lower()

            if "icon-arrowdown" in icon_class:
                is_spread_negative = True
                print("⬇️ 前日比: 下落（マイナス扱い）")
            else:
                is_spread_negative = False
                print("⬆️ 前日比: 上昇/変わらず")

        except Exception as e:
            msg = str(e).split("\n")[0].replace("Message: ", "")
            print(f"⚠️ 前日比アイコン未検出 ({msg})。")
            is_spread_negative = False

        # --- 2) データ取得 ---
        for name, selector in XPATH_SELECTORS.items():
            if name == "Spread_Icon":
                continue

            try:
                element = driver.find_element(By.XPATH, selector)
                text = (element.text or "").strip()

                # VNIndex (float)
                if name == "VNIndex":
                    data_row_internal["VNIndex"] = parse_float(text)
                    print(f"   -> {LOG_LABEL['VNIndex']}: {data_row_internal['VNIndex']}")
                    continue

                # Spread / Spread% (float,float)
                if name == "Spread":
                    raw_spread = None
                    raw_spread_percent = None

                    # Example patterns: "12.34 0.56%" or "12.34 / 0.56%"
                    match = re.search(r"([-+]?\d[\d\.,]*)(?:\s+|/)([-+]?\d[\d\.,]*%)", text)
                    if match:
                        raw_spread = parse_float(match.group(1))
                        raw_spread_percent = parse_float(match.group(2))
                    else:
                        # fallback: find first number and first percent
                        nums = re.findall(r"[-+]?\d+(?:\.\d+)?", _clean_number_text(text))
                        perc = re.findall(r"[-+]?\d+(?:\.\d+)?(?=%)", text)
                        raw_spread = float(nums[0]) if nums else None
                        raw_spread_percent = float(perc[0]) if perc else None

                    if is_spread_negative:
                        if raw_spread is not None:
                            raw_spread = -abs(raw_spread)
                        if raw_spread_percent is not None:
                            raw_spread_percent = -abs(raw_spread_percent)

                    data_row_internal["Spread"] = raw_spread
                    data_row_internal["Spread%"] = raw_spread_percent

                    print(f"   -> {LOG_LABEL['Spread']}: {data_row_internal['Spread']}")
                    print(f"   -> {LOG_LABEL['Spread%']}: {data_row_internal['Spread%']}")
                    continue

                # Value (float, in 'tỷ' unit)
                if name == "Value":
                    data_row_internal["Value"] = parse_value_ty(text)
                    print(f"   -> {LOG_LABEL['Value']}: {data_row_internal['Value']}")
                    continue

                # Volume / CP_* (int)
                if name in ("Volume", "CP_Tang", "CP_Giam", "CP_KhongDoi"):
                    data_row_internal[name] = parse_int(text)
                    label = LOG_LABEL.get(name, name)
                    print(f"   -> {label}: {data_row_internal[name]}")
                    continue

                # fallback
                data_row_internal[name] = text

            except Exception as e:
                msg = str(e).split("\n")[0].replace("Message: ", "")
                label = LOG_LABEL.get(name, name)
                print(f"❌ 要素未検出: {label} | 詳細: {msg}")
                # keep None for numeric fields
                if name == "Spread":
                    data_row_internal["Spread"] = None
                    data_row_internal["Spread%"] = None
                else:
                    if name in data_row_internal:
                        data_row_internal[name] = None

        # --- 3) 重複チェック（numeric）---
        last_data_numeric = get_last_excel_data_numeric()

        # 現在データ（比較用）を “日本語カラム名” に変換
        current_data_jp = {}
        for internal_key in COMPARE_COLUMNS_INTERNAL:
            jp_col = COLUMN_JP[internal_key]
            current_data_jp[jp_col] = data_row_internal.get(internal_key, None)

        if is_duplicate_numeric(current_data_jp, last_data_numeric, float_tol=1e-6):
            print("\n=======================================================")
            print("🚫 現在データは Excel の最終行と同一です。")
            print("➡️ 既に最新データが保存されています（取引終了の可能性あり）。")
            print("=======================================================\n")
            return

        print("\n=======================================================")
        print("✅ 新しいデータを取得しました！")
        print("➡️ 今回取得したデータを Excel に追記します。")
        print("=======================================================\n")

        # --- 4) Excel へ保存（日本語ヘッダー、numericで保存）---
        trading_date = get_trading_date()

        data_row_jp = {
            COLUMN_JP["ThoiGian"]: trading_date,
            COLUMN_JP["VNIndex"]: data_row_internal.get("VNIndex"),
            COLUMN_JP["Spread"]: data_row_internal.get("Spread"),
            COLUMN_JP["Spread%"]: data_row_internal.get("Spread%"),
            COLUMN_JP["Value"]: data_row_internal.get("Value"),
            COLUMN_JP["Volume"]: data_row_internal.get("Volume"),
            COLUMN_JP["CP_Tang"]: data_row_internal.get("CP_Tang"),
            COLUMN_JP["CP_Giam"]: data_row_internal.get("CP_Giam"),
            COLUMN_JP["CP_KhongDoi"]: data_row_internal.get("CP_KhongDoi"),
        }

        df_out = pd.DataFrame([data_row_jp])[FINAL_COLUMN_ORDER_JP]

        print(f"💾 Excel に保存: {EXCEL_FILE_PATH}")
        file_exists = os.path.isfile(EXCEL_FILE_PATH)

        if file_exists:
            try:
                book = load_workbook(EXCEL_FILE_PATH)
                sheet = book.active

                # 既存ヘッダー確認（日本語じゃなければ作り直し）
                existing_header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                if existing_header != FINAL_COLUMN_ORDER_JP:
                    raise ValueError("既存Excelのヘッダーが日本語カラムと一致しません（作り直しを実行）。")

                with pd.ExcelWriter(
                    EXCEL_FILE_PATH,
                    engine="openpyxl",
                    mode="a",
                    if_sheet_exists="overlay"
                ) as writer:
                    sheet2 = writer.book.active
                    start_row = sheet2.max_row
                    df_out.to_excel(
                        writer,
                        sheet_name=sheet2.title,
                        startrow=start_row,
                        index=False,
                        header=False
                    )

            except Exception as e:
                print(f"⚠️ 追記に失敗: {e}")
                print("➡️ 日本語カラムで新規作成（上書き）します。")
                df_out.to_excel(EXCEL_FILE_PATH, index=False, header=True, engine="openpyxl")
        else:
            df_out.to_excel(EXCEL_FILE_PATH, index=False, header=True, engine="openpyxl")

        print("🎉 保存完了！（numericで保存されています）")

    except Exception as e:
        print(f"❌ 全体処理エラー: {e}")

    finally:
        if driver:
            driver.quit()
            print("🔒 ブラウザを終了しました。")


# =========================================================================
# MAIN
# =========================================================================

if __name__ == "__main__":
    get_market_data_and_save()
